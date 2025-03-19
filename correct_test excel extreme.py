# 用于极值剪枝效率，求时间
from openpyxl import Workbook
from openpyxl import load_workbook
from correct_test import * 
# 创建一个工作簿
filename = '不加剪枝.xlsx'
sheetname ='Sheet1'
# 加载现有的Excel文件
wb = load_workbook(filename)
# 选择活动工作表或指定工作表
ws = wb[sheetname]  # 或者使用 ws = wb['Sheet1'] 来指定工作表
# 将数值写入单元格
latticeType=1
dimension=100
q=128
blksize=62
type=3
for dimension in range(100, 131):  # 假设我们写入10行数据
    print("dimension:",dimension)
    # if (selectmode(latticeType,dimension,blksize,q)):
    #     _ = ws.cell(row=dimension, column=1, value=blksize)
    #     _ = ws.cell(row=dimension, column=2, value=blksize)
    # else:
    #     blksize+=1
    #     simension-=1
    corre,times=selectmode(latticeType,type,dimension,blksize,q)
    while (not corre):
        if(blksize<=100):
            blksize+=1
        corre,times=selectmode(latticeType,type,dimension,blksize,q)
    else:
        _ = ws.cell(row=dimension+1, column=1, value=dimension)
        _ = ws.cell(row=dimension+1, column=2, value=blksize)
        _ = ws.cell(row=dimension+1, column=3, value=times)
        wb.save(filename)

# 保存工作簿到文件
wb.save(filename)
