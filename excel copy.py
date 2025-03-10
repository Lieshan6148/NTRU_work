from openpyxl import Workbook
from openpyxl import load_workbook
from datagen import *
# 创建一个工作簿
filename = '数值表.xlsx'

# 加载现有的Excel文件
wb = load_workbook(filename)

# 选择活动工作表或指定工作表
ws = wb['Sheet1']  # 或者使用 ws = wb['Sheet1'] 来指定工作表
# 将数值写入单元格
for row in range(1, 21):  # 假设我们写入10行数据
    print("dimension:",row*5) 
    x,y,z=timecomp(row*5)
    _ = ws.cell(row=row+1, column=2, value=x)
    _ = ws.cell(row=row+1, column=3, value=y)
    _ = ws.cell(row=row+1, column=4, value=z)

# 保存工作簿到文件
wb.save("数值表.xlsx")
